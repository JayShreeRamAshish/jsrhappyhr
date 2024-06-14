import streamlit as st
from sqlalchemy import create_engine,inspect,Column,Integer,String,DateTime,Boolean,ForeignKey
from sqlalchemy.orm import sessionmaker,relationship, declarative_base
from setup_db import User,Employee,Base,SurveyTemplate,SurveyRollout,Branch,Department,Designation,Grade,SalaryStructure,SalaryHead
import pandas as pd
from io import BytesIO
import xlsxwriter
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.workbook.protection import WorkbookProtection
import matplotlib.pyplot as plt
from datetime import datetime
import json
from email_validator import validate_email, EmailNotValidError
import smtplib
from fpdf import FPDF
import os
from dotenv import load_dotenv

st.set_page_config(page_title="HappyHR",page_icon="🧊",layout="wide",initial_sidebar_state="expanded")

st.markdown('<style>{}</style>'.format(open('style.css').read()), unsafe_allow_html=True)

# Load environment variables
load_dotenv()

# Create an engine
engine = create_engine('sqlite:///app.db')
Base.metadata.create_all(engine)
Session = sessionmaker(bind=engine)

class SalaryHeadManagement:
    def __init__(self, session):
        self.session = session

    def create_salary_head(self, head_code, head_name, head_description, calculate_on_weo, calculate_on_phy, calculate_on_attendance):
        new_head = SalaryHead(
            head_code=head_code,
            head_name=head_name,
            head_description=head_description,
            calculate_on_weo=calculate_on_weo,
            calculate_on_phy=calculate_on_phy,
            calculate_on_attendance=calculate_on_attendance
        )
        self.session.add(new_head)
        self.session.commit()

    def view_salary_heads(self):
        return self.session.query(SalaryHead).all()

class SalaryStructureManagement:
    def __init__(self, session):
        self.session = session

    def create_salary_structure(self, employee_id, salary_head_ids):
        for head_id in salary_head_ids:
            new_structure = SalaryStructure(employee_id=employee_id, salary_head_id=head_id, value=0)
            self.session.add(new_structure)
        self.session.commit()

    def get_salary_structure(self, employee_id):
        return self.session.query(SalaryStructure).filter_by(employee_id=employee_id).all()

    def modify_salary_structure(self, employee_id, salary_head_ids):
        existing_structures = self.get_salary_structure(employee_id)
        for structure in existing_structures:
            self.session.delete(structure)
        self.session.commit()
        self.create_salary_structure(employee_id, salary_head_ids)

    def update_salary_structure(self, employee_id, head_values):
        structures = self.get_salary_structure(employee_id)
        for structure in structures:
            if structure.salary_head_id in head_values:
                structure.value = head_values[structure.salary_head_id]
        self.session.commit()

    def export_salary_structure_to_excel(self):
        structures = self.session.query(SalaryStructure).all()
        data = [{
            "Employee ID": structure.employee_id,
            "Salary Head ID": structure.salary_head_id,
            "Value": structure.value
        } for structure in structures]
        df = pd.DataFrame(data)
        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            writer.save()
        output.seek(0)
        return output

    def export_salary_structure_to_pdf(self):
        structures = self.session.query(SalaryStructure).all()
        data = [{
            "Employee ID": structure.employee_id,
            "Salary Head ID": structure.salary_head_id,
            "Value": structure.value
        } for structure in structures]
        df = pd.DataFrame(data)

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        
        col_width = pdf.w / 4.5
        row_height = pdf.font_size
        
        for row in df.values:
            for item in row:
                pdf.cell(col_width, row_height*1.25, str(item), border=1)
            pdf.ln(row_height*1.25)
        
        pdf_output = BytesIO()
        pdf.output(pdf_output)
        pdf_output.seek(0)
        return pdf_output

    def import_salary_structure_from_excel(self, file):
        df = pd.read_excel(file)
        for _, row in df.iterrows():
            new_structure = SalaryStructure(
                employee_id=row['Employee ID'],
                salary_head_id=row['Salary Head ID'],
                value=row['Value']
            )
            self.session.add(new_structure)
        self.session.commit()
        

class Masterddl:
    def __init__(self, session):
        self.session = session

    def create_Branch(self,branch_code,branch_name):
        new_branch=Branch(branch_code=branch_code,branch_name=branch_name)
        self.session.add(new_branch)
        self.session.commit()
        
    def create_Department(self,session):
        new_department=Department(department_code=department_code,department_name=department_name)
        self.session.add(new_department)
        self.session.commit()
        
    def create_Designation(self,session):
        new_designation=Designation(designation_code=designation_code,designation_name=designation_name)
        self.session.add(new_designation)
        self.session.commit()
        
    def create_Grade(self,session):
        pass
    
    def View_Branch(self,session):
        pass
    def VieweDepartment(self,session):
        pass
    def ViewDesignation(self,session):
        pass
    def VieweGrade(self,session):
        pass
    
    def EditBranch(self,session):
        pass
    def EditeDepartment(self,session):
        pass
    def EditeDesignation(self,session):
        pass
    def EditGrade(self,session):
        pass
    
            
class SQLPlayGround:
    def __init__(self, engine):
        self.engine = engine
    
    def execute_query(self, query):
        if not query.strip():
            return pd.DataFrame()  # Return an empty DataFrame if the query is null
        try:
            with self.engine.connect() as connection:
                result = pd.read_sql_query(query, connection)
                return result
        except Exception as e:
            st.error(f"An error occurred: {e}")
            return pd.DataFrame()  # Return an empty DataFrame in case of error

    def export_to_excel(self, df, password=None):
        try:
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')
            processed_data = output.getvalue()
            output.seek(0)  # Rewind the buffer to the beginning after writing
            return processed_data
        except Exception as e:
            st.error(f"An error occurred while exporting to Excel: {e}")
            return None

    def export_to_pdf(self, df):
        try:
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=12)
            for i in range(len(df)):
                pdf.cell(200, 10, txt=str(df.iloc[i].to_list()), ln=True)
            output = BytesIO()
            pdf.output(dest='S').encode('latin1')  # Output to string and encode it properly
            output.write(pdf.output(dest='S').encode('latin1'))
            output.seek(0)  # Rewind the buffer to the beginning after writing
            return output.getvalue()
        except Exception as e:
            st.error(f"An error occurred while exporting to PDF: {e}")
            return None

    def generate_graph(self, df, graph_type, title):
        plt.figure()
        if graph_type == 'Line Chart':
            df.plot(kind='line', title=title)
        elif graph_type == 'Bar Chart':
            df.plot(kind='bar', title=title)
        elif graph_type == 'Area Chart':
            df.plot(kind='area', title=title)
        elif graph_type == 'Scatter Chart':
            df.plot(kind='scatter', x=df.columns[0], y=df.columns[1], title=title)
        elif graph_type == 'Histogram':
            df.plot(kind='hist', title=title)
        elif graph_type == 'Box Plot':
            df.plot(kind='box', title=title)
        elif graph_type == 'Pie Chart':
            df.plot(kind='pie', y=df.columns[0], title=title)
        plt.tight_layout()
        output = BytesIO()
        plt.savefig(output, format='png')
        plt.close()
        output.seek(0)
        return output

def get_tables_and_fields(engine):
        inspector = inspect(engine)
        tables_info = {}
        for table_name in inspector.get_table_names():
            columns = [column['name'] for column in inspector.get_columns(table_name)]
            tables_info[table_name] = columns
        return tables_info


# Assuming you have already created an SQLAlchemy engine object
# engine = create_engine('your_database_connection_string')

# Streamlit part
# User management
class UserManagement:
    def __init__(self, session):
        self.session = session   
    
    def create_user(self, username, password):
        new_user = User(username=username, password=password)
        self.session.add(new_user)
        self.session.commit()
        
    def get_employee_by_id(self, emp_id):
        return self.session.query(Employee).filter_by(id=emp_id).first()

    def view_users(self):
        return self.session.query(User).all()
    
    def download_employees_as_excel(self):
        employees = self.view_employees()
        df = pd.DataFrame([(emp.id, emp.emcode, emp.emfname, emp.emlname, emp.employee_fathername,
                            emp.emdob, emp.emdoj, emp.emdoc, emp.embranchpid, emp.emdepartmentpid,
                            emp.emdesignationpid, emp.emgradepid, emp.empan, emp.emaadhar, emp.emuan)
                           for emp in employees],
                          columns=['ID', 'Employee Code', 'First Name', 'Last Name', 'Father Name',
                                   'DOB', 'DOJ', 'DOC', 'Branch', 'Department', 'Designation', 'Grade', 'PAN', 'Aadhar', 'UAN']) 
        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Employees')
            writer.save()
        processed_data = output.getvalue()
        return processed_data

    def download_employees_as_pdf(self):
        employees = self.view_employees()
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        for emp in employees:
            pdf.cell(200, 10, txt=f"ID: {emp.id}, Code: {emp.emcode}, Name: {emp.emfname} {emp.emlname}, Father: {emp.employee_fathername}, DOB: {emp.emdob}, DOJ: {emp.emdoj}, DOC: {emp.emdoc}, Branch: {emp.embranchpid}, Dept: {emp.emdepartmentpid}, Designation: {emp.emdesignationpid}, Grade: {emp.emgradepid}, PAN: {emp.empan}, Aadhar: {emp.emaadhar}, UAN: {emp.emuan}", ln=True)
        return pdf.output(dest='S').encode('latin1')
    

    def modify_user(self, user_id, username, password):
        user = self.session.query(User).filter_by(id=user_id).first()
        user.username = username
        user.password = password
        self.session.commit()

# Employee Central
class EmployeeManagement:
    def __init__(self, session):
        self.session = session

    def create_employee(self,emcode,emfname,emlname,employee_fathername,emdob,emdoj,emdoc,embranchpid,emdepartmentpid,emdesignationpid,emgradepid,empan,emaadhar,emuan):
        new_employee = Employee(
            emcode = emcode,
            emfname = emfname,
            emlname = emlname,
            employee_fathername = employee_fathername,
            emdob = emdob,
            emdoj = emdoj,
            emdoc = emdoc,
            embranchpid = embranchpid,
            emdepartmentpid = emdepartmentpid,
            emdesignationpid = emdesignationpid,
            emgradepid = emgradepid,
            empan = empan,
            emaadhar = emaadhar,
            emuan = emuan)
        self.session.add(new_employee)
        self.session.commit()

    def view_employees(self):
        return self.session.query(Employee).all()
    
    def get_employee_by_id(self,emp_id):
        return self.session.query(Employee).filter_by(id=emp_id).first()
    
    def modify_employee(self, emp_id, emcode, emfname, emlname, employee_fathername, emdob, emdoj, emdoc, embranchpid, emdepartmentpid, emdesignationpid, emgradepid, empan, emaadhar, emuan):
        employee = self.session.query(Employee).filter_by(id=emp_id).first()
        employee.emcode = emcode
        employee.emfname = emfname
        employee.emlname = emlname
        employee.employee_fathername = employee_fathername
        employee.emdob = emdob
        employee.emdoj = emdoj
        employee.emdoc = emdoc
        employee.embranchpid = embranchpid
        employee.emdepartmentpid = emdepartmentpid
        employee.emdesignationpid = emdesignationpid
        employee.emgradepid = emgradepid
        employee.empan = empan
        employee.emaadhar = emaadhar
        employee.emuan = emuan
        self.session.commit()
        
    def download_employees_as_excel(self):
        employees = self.view_employees()
        df = pd.DataFrame([(emp.id, emp.emcode, emp.emfname, emp.emlname, emp.employee_fathername,
                            emp.emdob, emp.emdoj, emp.emdoc, emp.embranchpid, emp.emdepartmentpid,
                            emp.emdesignationpid, emp.emgradepid, emp.empan, emp.emaadhar, emp.emuan)
                           for emp in employees],
                          columns=['ID', 'Employee Code', 'First Name', 'Last Name', 'Father Name',
                                   'DOB', 'DOJ', 'DOC', 'Branch', 'Department', 'Designation', 'Grade', 'PAN', 'Aadhar', 'UAN'])
        output = BytesIO()
        writer = pd.ExcelWriter(output, engine='xlsxwriter')
        df.to_excel(writer, index=False, sheet_name='Employees')
        writer.save()
        processed_data = output.getvalue()
        return processed_data

    def download_employees_as_pdf(self):
        employees = self.view_employees()
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        for emp in employees:
            pdf.cell(200, 10, txt=f"ID: {emp.id}, Code: {emp.emcode}, Name: {emp.emfname} {emp.emlname}, Father: {emp.employee_fathername}, DOB: {emp.emdob}, DOJ: {emp.emdoj}, DOC: {emp.emdoc}, Branch: {emp.embranchpid}, Dept: {emp.emdepartmentpid}, Designation: {emp.emdesignationpid}, Grade: {emp.emgradepid}, PAN: {emp.empan}, Aadhar: {emp.emaadhar}, UAN: {emp.emuan}", ln=True)
        return pdf.output(dest='S').encode('latin1')
    

# Login authentication
def login(session, username, password):
    return session.query(User).filter_by(username=username, password=password).first()

# Main Streamlit application
def main():

    # Authentication
    if 'logged_in' not in st.session_state:
        st.session_state.logged_in = False

    if not st.session_state.logged_in:
        st.sidebar.header("Happy-HR")
        username = st.sidebar.text_input("Username")
        password = st.sidebar.text_input("Password", type="password")
        if st.sidebar.button("Login"):
            session = Session()
            user = login(session, username, password)
            if user:
                st.session_state.logged_in = True
                st.session_state.username = username
                st.success("Logged in successfully")
            else:
                st.error("Invalid username or password")
            session.close()
        return

    st.sidebar.header("Happy-HR")
    
    menu=st.sidebar.selectbox("HappyHR",["Dashboard","Employee Central","Attendance","Payroll","Recruitment","PMS","eLearning","Data Central","Manage Users","Developer","Survey Builder"])

    session = Session()
    
    if menu=="Employee Central":
        #st.title("Employee Database Central")
        sub_menu = st.sidebar.selectbox("Select Submenu", ["View Employee","New Employee","Modify Employee","Master"])
        emp_mgmt = EmployeeManagement(session)
        col1,col2=st.columns([1,2])

        if sub_menu == "New Employee":
            with col1:
                emcode = st.text_input("Employee Code")
                role = st.text_input("Role")
                emfname = st.text_input("First Name")
                emlname = st.text_input("Last Name")
                employee_fathername = st.text_input("Father Name")
                emdob = st.text_input("Date of Birth")
                emdoj = st.text_input("Date of Joining")
                emdoc = st.text_input("Date of Confirmation")
            with col2:
                embranchpid = st.text_input("Branch")
                emdepartmentpid = st.text_input("Department")
                emdesignationpid = st.text_input("Designation")
                emgradepid = st.text_input("Grade")
                empan = st.text_input("PAN")
                emaadhar = st.text_input("Aadhar")
                emuan = st.text_input("PF UAN")            
            if st.button("Create Employee"):
                emp_mgmt.create_employee(emcode,emfname,emlname,employee_fathername,emdob,emdoj,emdoc,embranchpid,emdepartmentpid,emdesignationpid,emgradepid,empan,emaadhar,emuan)
                st.success("Employee created successfully")
                
        elif sub_menu=="View Employee":
            emp_mgmt = EmployeeManagement(session)
            #st.title("Employee Database")
            st.subheader("Employees Database",divider=True)
            employees = emp_mgmt.view_employees()
            data = [{
            "ID": emp.id,
            "Employee Code": emp.emcode,
            "First Name": emp.emfname,
            "Last Name": emp.emlname,
            "Father Name": emp.employee_fathername,
            "DOB": emp.emdob,
            "DOJ": emp.emdoj,
            "DOC": emp.emdoc,
            "Branch": emp.embranchpid,
            "Department": emp.emdepartmentpid,
            "Designation": emp.emdesignationpid,
            "Grade": emp.emgradepid,
            "PAN": emp.empan,
            "Aadhar": emp.emaadhar,
            "UAN": emp.emuan,
            } for emp in employees]
            df = pd.DataFrame(data)
            st.dataframe(df,use_container_width=True,hide_index=True,selection_mode="single-row")

            st.subheader("",divider=True)
            col1, col2 = st.columns(2)
            with col1:
                if st.button("Download as Excel"):
                    excel_data = emp_mgmt.download_employees_as_excel()
                    st.download_button(
                    label="Download Excel file",
                    data=excel_data,
                    file_name='employees.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

            with col2:
                if st.button("Download as PDF"):
                    pdf_data = emp_mgmt.download_employees_as_pdf()
                    st.download_button(
                    label="Download PDF file",
                    data=pdf_data,
                    file_name='employees.pdf',
                    mime='application/pdf')
                    session.close()
    
        elif sub_menu == "Modify Employee":
            emp_id = st.number_input("Employee ID", min_value=1, step=1)
            if st.button("Load Employee"):
                employee = emp_mgmt.get_employee_by_id(emp_id)
                if employee:
                    emcode = st.text_input("Employee Code", value=employee.emcode)
                    emfname = st.text_input("First Name", value=employee.emfname)
                    emlname = st.text_input("Last Name", value=employee.emlname)
                    employee_fathername = st.text_input("Father Name", value=employee.employee_fathername)
                    emdob = st.text_input("Date of Birth", value=employee.emdob)
                    emdoj = st.text_input("Date of Joining", value=employee.emdoj)
                    emdoc = st.text_input("Date of Confirmation", value=employee.emdoc)
                    embranchpid = st.text_input("Branch", value=employee.embranchpid)
                    emdepartmentpid = st.text_input("Department", value=employee.emdepartmentpid)
                    emdesignationpid = st.text_input("Designation", value=employee.emdesignationpid)
                    emgradepid = st.text_input("Grade", value=employee.emgradepid)
                    empan = st.text_input("PAN", value=employee.empan)
                    emaadhar = st.text_input("Aadhar", value=employee.emaadhar)
                    emuan = st.text_input("PF UAN", value=employee.emuan)
                    
                    if st.button("Modify Employee"):
                        emp_mgmt.modify_employee(emp_id,emcode,emfname,emlname,employee_fathername,emdob,emdoj,emdoc,embranchpid, emdepartmentpid, emdesignationpid,emgradepid,empan,emaadhar,emuan)
                        st.success("Employee modified successfully")
                    else:
                        st.error("Employee not found")
        elif sub_menu=="Master":
                col1,col2,col3,col4=st.columns(4)
                mast_ddl=Masterddl(session)
                with col1:
                    ddlmenu=st.selectbox("Drop Down List",["Create New Branch","Create New Deparment","Create New Designation","Create New Grade"])
                    
                    if ddlmenu=="Create New Branch":
                        st.write("Create New Branch")
                        branch_code=st.text_input("Branch Code")
                        branch_name=st.text_input("Branch NAme")
                        if st.button("Create Branch"):
                            mast_ddl.create_Branch(branch_code,branch_name)
                            st.write("Branch Created Sucessfully")
                            
                    if ddlmenu=="Create New Deparment":
                        st.write("Create New Deparment")
                        department_code=st.text_input("Department Code")
                        department_name=st.text_input("Department Name")
                        if st.button("Create Department"):
                            mast_ddl.create_Branch(department_code,department_name)
                            st.write("Department Created Sucessfully")
                            
                    if ddlmenu=="Create New Designation":
                        designation_code=st.text_input("Designation Code")
                        designation_name=st.text_input("Designation Name")
                        if st.button("Create Designation"):
                            mast_ddl.create_Designation(designation_code,designation_name)
                            st.write("Designation Created Sucessfully")
                            
                    if ddlmenu=="Create New Grade":
                        st.write("Create New Grade")
                        grade_code=st.text_input("Grade Code")
                        grade_name=st.text_input("Grade Name")
                        if st.button("Create Grade"):
                            mast_ddl.create_Grade(grade_code,grade_name)
                            st.write("Grade Created Sucessfuly")
                                
                        
                    
                with col2:
                    st.write("Create New Department")
                    
                        
                    st.write("View Department")
                    st.write("Edit Department")
                    
                with col3:
                    st.write("Create New Designation")
                    st.write("View Designation")
                    st.write("Edit Designation")
                    
                with col4:
                    st.write("Create New Grade")
                    st.write("View Grade")
                    st.write("Edit Grade")
                    
    elif menu == "Manage Users":
        st.subheader("User Management")
        sub_menu = st.sidebar.selectbox("Select Submenu", ["Create User", "View Users", "Modify User"])
        user_mgmt = UserManagement(session)
        if sub_menu == "Create User":
            username = st.text_input("Username")
            password = st.text_input("Password", type="password")
            
            if st.button("Create User"):
                    user_mgmt.create_user(username, password)
                    st.success("User created successfully")

        elif sub_menu == "View Users":
                users = user_mgmt.view_users()
                st.write(users)

        elif sub_menu == "Modify User":
                user_id = st.number_input("User ID", min_value=1)
                username = st.text_input("New Username")
                password = st.text_input("New Password", type="password")
                if st.button("Modify User"):
                    user_mgmt.modify_user(user_id, username, password)
                    st.success("User modified successfully")
                    
    elif menu=="Payroll":
        st.subheader("Salary Stucture")
        salary_head_mgmt = SalaryHeadManagement(session)
        salary_structure_mgmt = SalaryStructureManagement(session)
        sub_menu=st.sidebar.selectbox("Select Menu", ["Create Salary Head", "View Salary Heads", "Salary Structure", "View and Update Employee Salary Structure", "Import/Export Salary Structure"])
        st.title("Salary Structure Management")
        existing_structure = salary_structure_mgmt.get_salary_structure(employee_id)
        existing_head_ids = [structure.salary_head_id for structure in existing_structure]
        

        if sub_menu=="Create Salary Head":
            st.subheader("Create Salary Head")
            head_code = st.text_input("Head Code")
            head_name = st.text_input("Head Name")
            head_description = st.text_area("Head Description")
            calculate_on_weo = st.checkbox("Calculate on WEO")
            calculate_on_phy = st.checkbox("Calculate on PHY")
            calculate_on_attendance = st.checkbox("Calculate on Attendance")
    
            if st.button("Create Salary Head"):
                salary_head_mgmt.create_salary_head(head_code, head_name, head_description, calculate_on_weo, calculate_on_phy, calculate_on_attendance)
                st.success("Salary Head created successfully")
        
        elif sub_menu == "View Salary Heads":
            st.subheader("View Salary Heads")
            salary_heads = salary_head_mgmt.view_salary_heads()
            data = [{
            "ID": head.id,
            "Head Code": head.head_code,
            "Head Name": head.head_name,
            "Head Description": head.head_description,
            "Calculate on WEO": head.calculate_on_weo,
            "Calculate on PHY": head.calculate_on_phy,
            "Calculate on Attendance": head.calculate_on_attendance
            } for head in salary_heads]
            df = pd.DataFrame(data)
            st.dataframe(df)
        
            st.subheader("Select Applicable Salary Heads")
            selected_head_ids = []
            employee_id = st.number_input("Employee ID", min_value=1)
            
            for head in salary_heads:
                checked = st.checkbox(head.head_name, value=head.id in existing_head_ids)
            if checked:
                selected_head_ids.append(head.id)

            if st.button("Save Salary Structure"):
                if existing_structure:
                    salary_structure_mgmt.modify_salary_structure(employee_id, selected_head_ids)
            else:
                salary_structure_mgmt.create_salary_structure(employee_id, selected_head_ids)
            st.success("Salary Structure saved successfully")

        elif sub_menu == "View and Update Employee Salary Structure":
            st.subheader("View and Update Employee Salary Structure")
            employee_id = st.number_input("Employee ID", min_value=1)
            if st.button("Load Employee Salary Structure"):
                structures = salary_structure_mgmt.get_salary_structure(employee_id)
                head_values = {structure.salary_head_id: structure.value for structure in structures}

            for structure in structures:
                new_value = st.number_input(f"Value for {structure.salary_head_id}", value=structure.value)
                head_values[structure.salary_head_id] = new_value

            if st.button("Update Salary Structure"):
                salary_structure_mgmt.update_salary_structure(employee_id, head_values)
                st.success("Salary Structure updated successfully")

        elif sub_menu == "Import/Export Salary Structure":
            st.subheader("Import/Export Salary Structure")

            st.subheader("Export")
            if st.button("Export to Excel"):
                df = salary_structure_mgmt.export_salary_structure_to_excel()
                st.download_button(label="Download Excel", data=df, file_name='salary_structure.xlsx', mime='application/vnd.ms-excel')

            if st.button("Export to PDF"):
                pdf_output = salary_structure_mgmt.export_salary_structure_to_pdf()
                st.download_button(label="Download PDF", data=pdf_output, file_name='salary_structure.pdf', mime='application/pdf')

            st.subheader("Import")
            uploaded_file = st.file_uploader("Choose an Excel file", type="xlsx")
            if uploaded_file is not None:
                salary_structure_mgmt.import_salary_structure_from_excel(uploaded_file)
                st.success("Salary Structure imported successfully")



# Streamlit part


    elif menu == "Developer":
        st.subheader("Support")
        sub_menu = st.sidebar.selectbox("Select Submenu", ["SQL Support", "Report Builder", "Dashboard Builder","Survey Builder"])
    
        if sub_menu == "SQL Support":
            st.subheader("SQL Playground")
            if 'report_templates' not in st.session_state:
                st.session_state['report_templates'] = []

            if 'graph_templates' not in st.session_state:
                st.session_state['graph_templates'] = []
        # Display tables and fields
            tables_info = get_tables_and_fields(engine)
            with st.sidebar:
                st.write("Tables and Fields")
                for table, columns in tables_info.items():
                    st.write(f"**{table}**")
                for column in columns:
                    st.write(f"- {column}")
        
        query = st.text_area("Write your SQL query here")
        
        if st.button("Execute Query"):
            sql_playground = SQLPlayGround(engine)  # Pass the engine to SQLPlayGround
            df = sql_playground.execute_query(query)
            st.write(df)

            if not df.empty:
                if st.checkbox("Save this result as a report template"):
                    report_name = st.text_input("Enter Report Name")
                    report_password = st.text_input("Enter Password to Protect Excel Report", type="password")
                    if st.button("Save Report"):
                        st.session_state['report_templates'].append((report_name, report_password, df))
                        st.success("Report template saved successfully")

                if st.checkbox("Design a graph based on the result"):
                    graph_title = st.text_input("Enter Graph Title")
                    graph_type = st.selectbox("Select Graph Type", ['Line Chart', 'Bar Chart', 'Area Chart', 'Scatter Chart', 'Histogram', 'Box Plot', 'Pie Chart'])
                    if st.button("Generate Graph Preview"):
                        graph = sql_playground.generate_graph(df, graph_type, graph_title)
                        st.image(graph, use_column_width=True)
                        if st.button("Save Graph"):
                            st.session_state['graph_templates'].append((graph_title, graph_type, graph))
                            st.success("Graph template saved successfully")

                excel_data = sql_playground.export_to_excel(df)
                if excel_data:
                    st.download_button(
                        label="Download Excel",
                        data=excel_data,
                        file_name='output.xlsx',
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                
                pdf_data = sql_playground.export_to_pdf(df)
                if pdf_data:
                    st.download_button(
                        label="Download PDF",
                        data=pdf_data,
                        file_name='output.pdf',
                        mime='application/pdf'
                    )

        elif sub_menu == "Report Builder":
            st.subheader("Report Builder")
            report_template_names = [name for name, _, _ in st.session_state['report_templates']]
            selected_report = None
            
            if report_template_names:
                selected_report = st.selectbox("Select a Report Template", report_template_names)
                excel_data = None
            if selected_report:
                report_password = st.text_input("Enter Password to Open Report", type="password")
                report_df = next(df for name, password, df in st.session_state['report_templates'] if name == selected_report)
                excel_data = SQLPlayGround(engine).export_to_excel(report_df, password=report_password)
            
            if excel_data:
                    st.download_button(
                    label="Download Report as Excel",
                    data=excel_data,
                    file_name=f'{selected_report}.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            else:
                st.write("No report templates available")
                report_name = None  # Initialize report_name with a default value
            if st.button("Save this result as report template"):
                report_name = st.text_input("Enter Report Name")
            if report_name and not report_name.strip():
                st.warning("Please enter a valid report name.")
            else:
                report_password = st.text_input("Enter Password to Protect Excel Report", type="password")
            if st.button("Save Report"):
                st.session_state['report_templates'].append((report_name, report_password, df))
                st.success("Report template saved successfully")
                
                
        elif sub_menu == "Dashboard":
            st.subheader("Dashboard")
            graph_template_names = [name for name, _, _ in st.session_state['graph_templates']]
            if graph_template_names:
                for graph_title, graph_type, graph in st.session_state['graph_templates']:
                    st.image(graph, caption=graph_title, use_column_width=True)
            else:
                st.write("No graphs available")
    elif menu=="Survey Builder":
        st.subheader("SUrvey Biolder")
        sub_menu = st.sidebar.selectbox("Select Submenu", ["Create Survey", "Rollout Survey", "Survey Results"])    
        if sub_menu == "Create Survey":
            create_survey_template()
        elif sub_menu == "Rollout Survey":
                rollout_survey()
        elif sub_menu == "Survey Results":
                survey_results()

    # Logout button
    if st.sidebar.button("Logout"):
        st.session_state.logged_in = False
        st.session_state.username = ""
        st.success("Logged out successfully")

    session.close()

if __name__ == '__main__':
    main()